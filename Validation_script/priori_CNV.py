#!/usr/bin/env python3

# usage: python3 priori_CNV.py --exomedepth SRR645629_sorted_md_exome_calls_ranked.csv --annotSV SRR645629_sorted_md_exome_calls_ranked.annotated.tsv --output SRR645629_CNV.xlsx
import pandas as pd
import argparse
from openpyxl.styles import PatternFill
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

def standardize_zygocity(value):
    """Convert various genotype formats to standardized Zygocity values"""
    if pd.isna(value):
        return "Nocall"
    
    value = str(value).strip()
    
    # Handle common cases first
    if value == '0/0':
        return "Hom_ref"
    elif value == '0/1':
        return "Het"
    elif value == '1/1':
        return "Hom"
    elif value == './.':
        return "Nocall"
    
    # Handle less common cases
    elif value in ['0|1', '1|0']:
        return "Het_phased"
    elif value == '1/2':
        return "Het_alt"
    
    # Fallback for unexpected values
    return value
def format_ranked_cnv(df):
    """Format the Ranked_CNV dataframe according to specifications"""
    # Define required columns for CNV_Position calculation
    required_columns = ['SV_type', 'SV_chrom', 'SV_start', 'SV_end', 'SV_length']
    # Create empty DataFrame with all required columns if input is empty
    if df.empty:
        all_columns = required_columns + [
            'Gene_name', 'Gene_count', 'CytoBand', 'Zygocity', 'P_gain_phen',
            'P_gain_source', 'P_loss_phen', 'P_loss_source', 'HI', 'TS',
            'DDD_HI_percent', 'OMIM_ID', 'GnomAD_pLI', 'BF', 'ACMG_classification'
        ]
        df = pd.DataFrame(columns=all_columns)
    # Ensure SV_length is numeric
    if 'SV_length' in df.columns:
        df['SV_length'] = pd.to_numeric(df['SV_length'], errors='coerce')    
    # Only proceed with CNV_Position calculation if required columns exist
    if all(col in df.columns for col in required_columns):
        # Create CNV_Position column
        df['CNV_Position'] = df['SV_type'] + '_' + df['SV_chrom'] + ':' + df['SV_start'].astype(str) + '-' + df['SV_end'].astype(str)
        # Format SV_length to kb (only if we have numeric values)
        if pd.api.types.is_numeric_dtype(df['SV_length']):
            df['CNV_length'] = (df['SV_length'] / 1000).round(1).astype(str) + 'kb'
        else:
            df['CNV_length'] = ''
    else:
        # If required columns are missing, create empty formatted columns
        df['CNV_Position'] = ''
        df['CNV_length'] = ''        

    
    # Rename columns
    df = df.rename(columns={
        'P_gain_phen': 'Pathogenic_Gain_Disease',
        'P_gain_source': 'Pathogenic_Gain_Source',
        'P_loss_phen': 'Pathogenic_Loss_Disease',
        'P_loss_source': 'Pathogenic_Loss_Source',
        'HI': 'ClinGene_HI_score',
        'TS': 'ClinGene_TS_score',
        'DDD_HI_percent': 'DECIPHER_HI_score',
        'GnomAD_pLI': 'pLI'
    })
    
    # Select and reorder columns
    columns = [
        'CNV_Position', 'CNV_length', 'Gene_name', 'Gene_count', 'CytoBand', 'Zygocity',
        'Pathogenic_Gain_Disease', 'Pathogenic_Gain_Source', 'Pathogenic_Loss_Disease',
        'Pathogenic_Loss_Source', 'OMIM_ID', 'ClinGene_HI_score', 'ClinGene_TS_score',
        'DECIPHER_HI_score', 'pLI', 'BF', 'ACMG_classification'
    ]
    
    # Only keep columns that exist in the dataframe
    existing_columns = [col for col in columns if col in df.columns]
    return df[existing_columns]

def apply_tab_color(workbook, sheet_name, color):
    """Apply color to the worksheet tab"""
    if sheet_name in workbook.sheetnames:
        workbook[sheet_name].sheet_properties.tabColor = color

def auto_adjust_columns(worksheet):
    """Auto-adjust column widths"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column_letter].width = adjusted_width

def main():
    # Set up argument parser
    parser = argparse.ArgumentParser(description='Merge and process genomic files')
    parser.add_argument('--exomedepth', required=True, help='First input file (CSV) with id and BF columns')
    parser.add_argument('--annotSV', required=True, help='Second input file (TSV) with SV columns and INFO')
    parser.add_argument('--output', required=True, help='Output TSV file path')
    parser.add_argument('--keep-info', action='store_true', help='Keep original INFO column')
    args = parser.parse_args()

    # Read input files
    df1 = pd.read_csv(args.exomedepth)
    df2 = pd.read_csv(args.annotSV, sep='\t')

    # Process exomedepth.csv to extract coordinates
    print("Processing coordinate matching...")
    df1[['chrom', 'coords']] = df1['id'].str.split(':', expand=True)
    df1[['start', 'end']] = df1['coords'].str.split('-', expand=True)
    df1['chrom'] = df1['chrom'].str.replace('chr', '')

    # Convert to same data types for comparison
    df1['start'] = df1['start'].astype(int)
    df1['end'] = df1['end'].astype(int)
    df2['SV_start'] = df2['SV_start'].astype(int)
    df2['SV_end'] = df2['SV_end'].astype(int)

    # Merge the dataframes to add BF column
    print("Merging files...")
    merged = pd.merge(
        df2,
        df1[['chrom', 'start', 'end', 'BF']],
        left_on=['SV_chrom', 'SV_start', 'SV_end'],
        right_on=['chrom', 'start', 'end'],
        how='left'
    ).drop(['chrom', 'start', 'end', 'ID', 'QUAL', 'FILTER', 'FORMAT', 'Closest_left', 'Closest_right', 'Tx', 'Tx_version', 'Tx_start', 'Tx_end', 'Exon_count', 'Location', 'Location2', 'Dist_nearest_SS', 'Nearest_SS_type', 'Intersect_start', 'Intersect_end', 'RE_gene', 'P_ins_phen', 'P_ins_hpo', 'P_ins_source', 'P_ins_coord', 'B_loss_AFmax', 'B_ins_source', 'B_ins_coord', 'B_ins_AFmax', 'B_inv_source', 'B_inv_coord', 'B_inv_AFmax', 'po_B_gain_allG_source', 'po_B_gain_allG_coord', 'po_B_gain_someG_source', 'po_B_gain_someG_coord', 'po_B_loss_allG_source', 'po_B_loss_allG_coord', 'po_B_loss_someG_source', 'po_B_loss_someG_coord', 'GC_content_left', 'GC_content_right', 'Repeat_coord_left', 'Repeat_type_left', 'Repeat_coord_right', 'Repeat_type_right', 'Gap_left', 'Gap_right', 'SegDup_left', 'SegDup_right', 'ENCODE_blacklist_left', 'ENCODE_blacklist_characteristics_left', 'ENCODE_blacklist_right', 'ENCODE_blacklist_characteristics_right', 'B_gain_AFmax', 'ACMG', 'ExAC_synZ', 'ExAC_misZ', 'NCBI_gene_ID'], axis=1)

    # Process INFO column to extract specific fields
    print("Processing INFO column...")
    def extract_info_fields(info):
        fields = {}
        for part in info.split(';'):
            if '=' in part:
                key, value = part.split('=', 1)
                if key in ['NEXONS', 'RATIO']:
                    fields[key] = value
        return pd.Series(fields)

    # Extract fields and add as new columns
    info_fields = merged['INFO'].apply(extract_info_fields)
   
    # Combine with main dataframe
    if args.keep_info:
        merged = pd.concat([merged, info_fields], axis=1)
    else:
        merged = pd.concat([merged.drop('INFO', axis=1), info_fields], axis=1)
    merged['NEXONS'] = merged['NEXONS'].astype(int)
    # Rename 10th column (index 9) to "Zygocity" and standardize values
    print("Standardizing Zygocity column...")
    col_name = merged.columns[9]
    merged = merged.rename(columns={col_name: "Zygocity"})
    merged['Zygocity'] = merged['Zygocity'].apply(standardize_zygocity)

    # Add ACMG classification tags
    print("Adding ACMG classification...")
    acmg_mapping = {
        '1': 'Benign',
        '2': 'Likely Benign',
        '3': 'VUS',
        '4': 'Likely Pathogenic',
        '5': 'Pathogenic'
    }

    # Convert HI/TS to numeric if columns exist
    for col in ['HI', 'TS']:
        if col in merged.columns:
            merged[col] = pd.to_numeric(merged[col], errors='coerce')

    # Convert ACMG_class to string and clean it
    merged['ACMG_class'] = merged['ACMG_class'].astype(str).str.strip()
    
    # Create new classification column
    merged['ACMG_classification'] = merged['ACMG_class'].map(acmg_mapping)
    
    # For any unmapped values, keep original
    merged['ACMG_classification'] = merged['ACMG_classification'].fillna(merged['ACMG_class'])

    # Create BF_ACMG filtered sheet - only numeric 3,4,5 values
    print("Creating BF_ACMG filtered sheet...")
    
    # Convert ACMG_class to string and strip whitespace for comparison
    merged['ACMG_class'] = merged['ACMG_class'].astype(str).str.strip()
    
    # Filter for numeric 3,4,5 only (not "full=3", etc.)
    acmg_numeric = merged['ACMG_class'].isin(['3', '4', '5'])

    # Combined filter for BF > 10 and numeric ACMG 3,4,5
    bf_acmg_conditions = (
        (merged['BF'].astype(float) > 10) & 
        acmg_numeric
    )
    bf_acmg_filtered = merged[bf_acmg_conditions]


    # Create pathogenic_CNV sheet if P_gain_coord or P_loss_coord has data
    pathogenic_cnv = None
    if 'P_gain_coord' in bf_acmg_filtered.columns or 'P_loss_coord' in bf_acmg_filtered.columns or 'po_P_gain_coord' in bf_acmg_filtered.columns or 'po_P_loss_coord' in bf_acmg_filtered.columns:
        print("Creating pathogenic_CNV sheet...")
        coord_condition = (
            bf_acmg_filtered['P_gain_coord'].notna() | 
            bf_acmg_filtered['P_loss_coord'].notna() |
            bf_acmg_filtered['po_P_gain_coord'].notna() |
            bf_acmg_filtered['po_P_loss_coord'].notna()
        )
        pathogenic_cnv = bf_acmg_filtered[coord_condition]

    # Create OMIM_CNV sheet if OMIM_morbid or OMIM_morbid_candidate contains "YES"
    omim_cnv = None
    if pathogenic_cnv is not None and not pathogenic_cnv.empty:
        omim_columns = ['OMIM_morbid', 'OMIM_morbid_candidate']
        existing_omim_cols = [col for col in omim_columns if col in pathogenic_cnv.columns]
        
        if existing_omim_cols:
            print("Creating OMIM_CNV sheet...")
            omim_condition = pathogenic_cnv[existing_omim_cols].apply(
                lambda x: x.str.upper().str.strip() == 'YES'
            ).any(axis=1)
            omim_cnv = pathogenic_cnv[omim_condition]
    
    # Create Ranked_CNV sheet for HI/TS scores 2 or 3 from OMIM_CNV (numeric comparison)
    ranked_cnv = None
    if omim_cnv is not None and not omim_cnv.empty:
        hi_ts_columns = ['HI', 'TS']
        existing_hi_ts_cols = [col for col in hi_ts_columns if col in omim_cnv.columns]
        
        if existing_hi_ts_cols:
            print("Creating Ranked_CNV sheet from OMIM_CNV...")
            # Numeric comparison for HI/TS scores
            hi_ts_condition = omim_cnv[existing_hi_ts_cols].apply(
                lambda x: x.isin([2, 3])
            ).any(axis=1)
            ranked_cnv = omim_cnv[hi_ts_condition]
            
            # Filter and sort by Gene_count if column exists
            if ranked_cnv is not None and not ranked_cnv.empty and 'Gene_count' in ranked_cnv.columns:
                # Remove rows with Gene_count < 10
                ranked_cnv = ranked_cnv[ranked_cnv['Gene_count'] > 10]
                print(f"Filtered to {len(ranked_cnv)} rows with Gene_count > 10")

                # Sort by Gene_count in descending order
                if not ranked_cnv.empty:
                    ranked_cnv = ranked_cnv.sort_values('Gene_count', ascending=False)
                    print("Sorted Ranked_CNV by Gene_count (high to low)")                


    # Save output to Excel with sheets
    print(f"Writing output to {args.output}...")
    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        # Write all sheets first
        merged.to_excel(writer, sheet_name='All_Data', index=False)
        bf_acmg_filtered.to_excel(writer, sheet_name='BF_ACMG', index=False)
        
        if pathogenic_cnv is not None and not pathogenic_cnv.empty:
            pathogenic_cnv.to_excel(writer, sheet_name='pathogenic_CNV', index=False)
            print(f"Added pathogenic_CNV sheet with {len(pathogenic_cnv)} rows")
        
        if omim_cnv is not None and not omim_cnv.empty:
            omim_cnv.to_excel(writer, sheet_name='OMIM_CNV', index=False)
            print(f"Added OMIM_CNV sheet with {len(omim_cnv)} rows")
        

        # Always create Ranked_CNV sheet, even if empty
        ranked_cnv_to_write = ranked_cnv if (ranked_cnv is not None and not ranked_cnv.empty) else pd.DataFrame()
        ranked_cnv_formatted = format_ranked_cnv(ranked_cnv_to_write)
        # Create the sheet with formatted columns
        ranked_cnv_formatted.to_excel(writer, sheet_name='Ranked_CNV', index=False)
        
        # Get the workbook and worksheet objects
        workbook = writer.book
        worksheet = workbook['Ranked_CNV']
        
        # Apply green color only to the tab
        apply_tab_color(workbook, 'Ranked_CNV', '92D050')  # Green color
        
        # Auto-adjust column widths
        auto_adjust_columns(worksheet)
        
        # Set Ranked_CNV as the active sheet
        workbook.active = workbook['Ranked_CNV']
        if not ranked_cnv.empty:
            print(f"Added formatted Ranked_CNV sheet with {len(ranked_cnv_formatted)} rows (green tab)")
        else:
            print("Created empty Ranked_CNV sheet (no variants met criteria: HI/TS 2-3 and Gene_count > 10)")
    
    print(f"Output saved with {len(merged)} rows in All_Data and {len(bf_acmg_filtered)} rows in BF_ACMG")
    print("Done!")                


if __name__ == '__main__':
    main()
